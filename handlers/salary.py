import re
from datetime import datetime

from aiogram import Bot, Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery

from database.general_db_functions import get_data_from_column, v_look_up_many, update_data_in_column
from database.user_table_functions import get_user_employee_code_from_db, get_user_state_from_db
from hidden.tokenfile import TOKEN_FOUR
from states import BossHere
from database.salary_table_functions import insert_dict_of_persons_to_database

import openpyxl

from keyboards.simple_keyboard import make_inline_many_keys_keyboard, make_inline_rows_keyboard, \
    make_inline_secret_many_keys_keyboard
from states import Registration

router = Router()
bot = Bot(TOKEN_FOUR)
TABLE_NAME = 'salary'

# Для поиска нужного листа мотивации ищем на каждом листе в диапазоне А1:С3 значение:
try_sheet_factor = 'Мотивация по не  учетным данным'

# Определение начального диапазона целевого листа для поиска нужных столбцов (код сотрудника, ЗП, премия и пр.)
start_cell = "A1"
end_cell = "I210"

# Получение численных значений для столбцов и строк
start_col = openpyxl.utils.column_index_from_string(start_cell[:1])
start_row = int(start_cell[1:])
end_col = openpyxl.utils.column_index_from_string(end_cell[:1])
end_row = int(end_cell[1:])

# Наименование столбцов, как они указаны в ячейке Excel-файла:
base_column = "Код."
target_column_01 = "Должность"
target_column_02 = "Ф.И.О."
target_column_03 = "Итог З/П"  # Оклад
target_column_04 = "Итог мотивация"  # Премия
target_column_05 = "Итог З\П+Бонус"  # Всего
target_column_06 = "Премия(компенсация отпуска)"
target_column_07 = "Вычеты ОС(форма/прочее)"
target_column_08 = "Вычеты ОС(Инвентаризация)"
target_column_09 = 'Вычеты-штрафы'
target_column_10 = 'Факт часы'
target_column_11 = 'Кол-во ошибок (примечание)'
target_column_12 = 'Сумма ошибки'
target_column_13 = 'Единый коэфф.'
target_column_14 = 'Дополнительные работы Н/Ч'
target_column_15 = 'Выдача '  # <-- пробел после слова - как в файле
target_column_16 = 'Доставки(Подготовка отгрузок)'
target_column_17 = 'Приемка'
target_column_18 = 'Размещение'
target_column_19 = 'Объем М3 кросс.'
target_column_20 = 'Сборка отгрузок'

target_columns = [base_column, target_column_01, target_column_02, target_column_03, target_column_04, target_column_05,
                  target_column_06, target_column_07, target_column_08, target_column_09, target_column_10,
                  target_column_11, target_column_12, target_column_13, target_column_14, target_column_15,
                  target_column_16, target_column_17, target_column_18, target_column_19, target_column_20]

# Так как в таблице наименования столбцов повторяются, а функция поиска нужного столбца возвращает только первый
# # найденный, применяем "сдвиг" для определенных столбцов:
offset_columns = [target_column_15, target_column_16, target_column_17,
                  target_column_18, target_column_19, target_column_20]
# Величина сдвига:
amount_of_indentation = 12

dict_of_persons = {}  # Глобальный словарь для формирования заливки
small_message_for_delete = []  # Глобальный список id сообщений-квитков для удалений
start_message_for_delete = []  # Глобальный список id сообщений для удаления в случае отмены табеля или его заливки


async def delete_some_messages(chat_id: int, numbers_of_message: list[int]) -> None:
    """Функция удаления сообщений в чате по message_id"""
    """Не забудь присвоить пустой список переменной, в случае передачи в эту функцию глобального списка"""
    for num in numbers_of_message:
        await bot.delete_message(chat_id=chat_id, message_id=num)


def search_salary_value(target_sheet, base_cell_name=base_column) -> None:
    """Функция создания и заполнения global dict_of_persons на основании данных в переданном листе target_sheet"""
    global dict_of_persons, base_column, target_columns
    # Ищем базовую ячейку (шапку базового столбца) в диапазоне start_cell:end_cell
    head_of_base_column = None
    for row in target_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if cell.value == base_cell_name:
                head_of_base_column = cell
                break
        if head_of_base_column:
            break
    # Если базовая ячейка (шапка базового столбца) найдена, считаем, что найден лист с таблицей, ищем целевые столбцы:
    if target_sheet:
        dict_of_persons = {}
        # Для каждого целевого столбца запускаем функцию переноса значений целевого столбца в global dict_of_persons
        for target in target_columns:
            search_values_of_one_target_column(
                base_column_head=head_of_base_column, target_sheet=target_sheet,
                target_column_name=target
            )

    # Добавляем значения базового столбца внутрь словаря
    for emp_no in dict_of_persons:
        dict_of_persons[emp_no][base_column] = emp_no

    # После полного формирования словаря, обнуляем значения строчных показателей для должностей с нестрочной мотивацией:
    positions_to_reset = ['Начальник склада', 'Заместитель начальника склада', 'Логист-претензионист',
                          'Супервизор', 'Старший кладовщик', 'Диспетчер', 'Старший оператор склада', 'Оператор склада']
    keys_to_reset = ['Выдача ', 'Доставки(Подготовка отгрузок)', 'Приемка', 'Размещение', 'Объем М3 кросс.',
                     'Сборка отгрузок']
    for key, data in dict_of_persons.items():
        position = data.get('Должность', '')
        if position in positions_to_reset:
            for key_to_reset in keys_to_reset:
                data[key_to_reset] = None


def forming_small_results_of_table() -> str:
    """
    Формирует краткий текст для подтверждения заливки руководителем.
    В данном случае, в формате: "Фамилия: итог ЗП"
    """
    global dict_of_persons
    text = ''
    for person_no in dict_of_persons:

        surname = str(dict_of_persons[person_no]["Ф.И.О."]).split(' ')[0]
        if len(surname) < 15:
            extra_space = 15 - len(surname)
            surname += '  ' * extra_space

        summ = dict_of_persons[person_no][target_column_05]
        emp_no = dict_of_persons[person_no][base_column]

        text += f'{emp_no} {surname}{summ:_} руб.\n'
    return text


@router.callback_query(F.data.in_(["Посмотреть квиток"]))
async def select_an_employee(callback: CallbackQuery) -> None:
    """Функция вывода клавиатуры с сотрудниками, находящимися на данный момент в global dict_of_persons
    Клавиатура предполагает генерацию калбеков, перехватываемых функцией forming_results_for_one_employee"""
    global small_message_for_delete

    await callback.answer()
    global dict_of_persons
    # Формируем лист с "Фамилия И(мя)" для каждого в global dict_of_persons
    names = [
        (
                (dict_of_persons[emp_no]["Ф.И.О."]).split(' ')[0]
                + " " +
                (dict_of_persons[emp_no]["Ф.И.О."]).split(' ')[1][:1]
        )
        for emp_no in dict_of_persons]

    # Удаляем предыдущие сообщения о квитках
    await delete_some_messages(chat_id=callback.from_user.id, numbers_of_message=small_message_for_delete)
    small_message_for_delete = []

    # Формируем объект message в переменную для извлечения message_id и вывода списка сотрудников
    message = await callback.message.answer(
        text='Выберите сотрудника',
        reply_markup=make_inline_many_keys_keyboard(names)
    )
    small_message_for_delete.append(message.message_id)


@router.callback_query(F.data.startswith("view_"))
async def forming_results_for_one_employee(callback: CallbackQuery):
    """Функция формирования из global dict_of_persons квитка для сотрудника, чья фамилия перехвачена в калбеке"""
    await callback.answer()
    global dict_of_persons, small_message_for_delete

    # Из каллбека извлекаем "Фамилия И(мя)":
    _, last_name = callback.data.split('_')

    # Задание тематических блоков и порядка в котором будут представлены данные:
    output_order_one = ['Ф.И.О.', 'Должность']
    output_order_two = ['Итог З/П', 'Итог мотивация', 'Итог З\\П+Бонус']
    output_order_three = ["Премия(компенсация отпуска)", "Вычеты ОС(форма/прочее)",
                          "Вычеты ОС(Инвентаризация)", 'Вычеты-штрафы', 'Дополнительные работы Н/Ч']
    output_order_four = ['Кол-во ошибок (примечание)', 'Сумма ошибки']
    output_order_five = ['Факт часы', 'Единый коэфф.']
    output_order_six = ['Выдача ', 'Доставки(Подготовка отгрузок)',
                        'Приемка', 'Размещение', 'Объем М3 кросс.', 'Сборка отгрузок']

    composed_text = ""
    for emp_id, emp_info in dict_of_persons.items():
        if (emp_info.get('Ф.И.О.').split(' ')[0] + " " + emp_info.get('Ф.И.О.').split(' ')[1][:1]) == last_name:

            composed_text += f"\nКод сотрудника: {emp_id}\n"
            for value in output_order_one:
                composed_text += f"{value}: {dict_of_persons[emp_id][value]}\n"
            composed_text += "-" * 45 + "\n"
            composed_text += f"ИТОГ ЗП: {dict_of_persons[emp_id][output_order_two[2]]} руб.\n"\
                             f"Оклад: {dict_of_persons[emp_id][output_order_two[0]]} руб.\n" \
                             f"Премия: {dict_of_persons[emp_id][output_order_two[1]]} руб.\n"
            composed_text += "-" * 45 + "\n"

            for value in output_order_three:
                if dict_of_persons[emp_id][value] is None:
                    text = "---"
                else:
                    text = f"{dict_of_persons[emp_id][value]} руб."
                composed_text += f"{value}: {text}\n"
            composed_text += "-" * 45 + "\n"

            if dict_of_persons[emp_id][output_order_four[0]] is None:
                composed_text += "Кол-во ошибок: ---\n"
            else:
                composed_text += f"Кол-во ошибок: {dict_of_persons[emp_id][output_order_four[0]]} " \
                                 f"(-{dict_of_persons[emp_id][output_order_four[1]]} руб.)\n"
            composed_text += "-" * 45 + "\n"

            for value in output_order_five:
                if dict_of_persons[emp_id][value] is None:
                    text = "---"
                else:
                    text = f"{dict_of_persons[emp_id][value]}"
                composed_text += f"{value}: {text}\n"
            composed_text += "-" * 45 + "\n"

            for value in output_order_six:
                if dict_of_persons[emp_id][value] is None:
                    text = "---"
                else:
                    text = f"{dict_of_persons[emp_id][value]}"
                composed_text += f"{value}: {text}\n"

            break  # Прекращаем цикл, так как нашли нужного сотрудника

    # Удаляем предыдущий квиток, если он был:
    await delete_some_messages(chat_id=callback.from_user.id, numbers_of_message=small_message_for_delete)
    small_message_for_delete = []

    # Формируем объект message в переменную для извлечения message_id и вывода квитка
    message = await callback.message.answer(
        text=composed_text,
        reply_markup=make_inline_rows_keyboard(["Выслать данные сотрудникам", "Посмотреть квиток", "Отменить"]))
    small_message_for_delete.append(message.message_id)


def search_values_of_one_target_column(base_column_head, target_sheet, target_column_name) -> None:
    """Функция извлечения данных из одного целевого столбца напротив непустых значений базового столбца (аналог ВПР)"""
    global dict_of_persons
    # Ищем целевой столбец.
    # Например, если базовый - "ID юзера", то целевым может быть - "Показатель" юзера по какому-то критерию
    target_column = None
    if base_column_head:
        # Проходим по ряду, в котором находятся истинные заголовки (строка 39):
        for col in target_sheet.iter_cols(min_row=39, max_row=39,
                                          min_col=base_column_head.column, max_col=target_sheet.max_column):

            # В каждом столбце всего одна строка, поэтому следующий цикл проходит один раз:
            for cell in col:
                target_column = False
                if cell.value == target_column_name:

                    # Если значение ячейки совпадает со столбцом, требующим сдвига, считаем det = 1, иначе det = 0:
                    if cell.value in offset_columns:
                        det = 1
                    else:
                        det = 0

                    # Проходим по строкам столбца col, добавляя каждую ячейку в кортеж target_column
                    target_column = tuple()
                    for cell_in_targ_col in target_sheet.iter_cols(min_row=base_column_head.row,
                                                                   max_row=target_sheet.max_row,
                                                                   min_col=col[0].column + amount_of_indentation * det,
                                                                   max_col=col[0].column + amount_of_indentation * det):
                        target_column += cell_in_targ_col

            # Прерываем цикл поиска целевого столбца, при первом найденном (ввиду дублирования наименования столбцов)
            if target_column:
                break

    # Если целевой столбец найден, выводим значения для непустых значений в базовом столбце
    if target_column:
        for cell in target_column:
            base_colum_value = target_sheet.cell(row=cell.row, column=base_column_head.column).value

            # Так же проверяем первые 4 символа (являются ли числами), чтобы отбросить строки-разделители
            if len(str(base_colum_value)) > 4 and str(base_colum_value)[:4].isdigit():

                # Добавляем значения в словарь
                if base_colum_value not in dict_of_persons:
                    dict_of_persons[base_colum_value] = {}
                cell_value = cell.value
                if type(cell_value) is float:
                    cell_value = round(cell_value, 2)
                dict_of_persons[base_colum_value][target_column_name] = cell_value


# Т.к. планируется, что бот будет перезагружаться, а значит, при заливке очередного табеля юзер вероятно будет
# "без статуса" ловим в роутере любое сообщение с файлом, и потом проверяем статус по базе:
@router.message(F.document.file_name.endswith('.xlsx'))
async def handle_excel_file(message: types.Document) -> None:
    """Функция анализа пойманного файла. При успешном результате выводятся краткие итоги по ЗП"""

    global start_message_for_delete

    # Смотрим статус в БД:
    name_of_current_state = get_user_state_from_db(str(message.chat.id))
    if name_of_current_state != 'employee_is_registered':
        await bot.send_message(
            chat_id=message.from_user.id,
            text='Вначале Вам необходимо пройти регистрацию /start'
        )
    else:

        # Получаем информацию о файле
        file_info = await bot.get_file(message.document.file_id)
        file = await bot.download_file(file_info.file_path)

        # Открываем файл с помощью openpyxl
        wb = openpyxl.load_workbook(file)

        # Перебираем все листы
        target_sheet = None
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Проверяем условие на каждом листе
            for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
                for cell in row:
                    if cell.value == try_sheet_factor:
                        # На первом листе, удовлетворяющем условию, останавливаем перебор:
                        target_sheet = sheet
                        break
            if target_sheet:
                break

        if target_sheet:
            # Используем target_sheet для дальнейших действий
            search_salary_value(target_sheet=target_sheet, base_cell_name=base_column)
            small_results = forming_small_results_of_table()
            if not small_results:
                small_results = 'Ничего не найдено'
                message_for_del = await bot.send_message(chat_id=message.from_user.id, text=small_results)
                start_message_for_delete.append(message_for_del.message_id)
            else:
                await delete_some_messages(chat_id=message.chat.id, numbers_of_message=[message.message_id])

                message_for_del = await bot.send_message(
                    chat_id=message.from_user.id,
                    text='В Вашем файле я обнаружил зарплатную таблицу.\nПроверьте итоговые суммы:')
                start_message_for_delete.append(message_for_del.message_id)

                message_for_del = await bot.send_message(chat_id=message.from_user.id, text=small_results)
                start_message_for_delete.append(message_for_del.message_id)

                message = await bot.send_message(
                    chat_id=message.from_user.id,
                    text='Вы можете посмотреть как будут выглядеть квитки для сотрудников.\n'
                         'После утверждения табеля к рассылке, просмотр квитков будет недоступен.',
                    reply_markup=make_inline_rows_keyboard(["Выслать данные сотрудникам", "Посмотреть квиток", "Отменить"]))
                small_message_for_delete.append(message.message_id)

        else:
            print("Лист не найден")

        # Закрываем скачанный файл
        wb.close()


@router.callback_query(F.data.startswith("Отменить"))
async def starting_to_create_password_report(callback: CallbackQuery):
    global dict_of_persons, start_message_for_delete, small_message_for_delete
    await callback.answer()

    await delete_some_messages(chat_id=callback.message.chat.id, numbers_of_message=small_message_for_delete)
    small_message_for_delete = []
    await delete_some_messages(chat_id=callback.message.chat.id, numbers_of_message=start_message_for_delete)
    start_message_for_delete = []
    dict_of_persons = {}


@router.callback_query(F.data.startswith("Выслать данные сотрудникам"))
async def starting_to_create_password_report(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    # Этот статус не сохраняется в БД, т.к. при перезагрузке бота потеряются данные из global dict_of_persons
    await state.set_state(BossHere.creating_a_secret_code)

    message_for_del = await callback.message.answer(text='Придумайте пароль для этой заливки. Каждая заливка '
                                                         'должна иметь уникальный пароль.\n'
                                                         'Зарегистрированные сотрудники смогут посмотреть свой квиток '
                                                         'только по этому паролю.')
    small_message_for_delete.append(message_for_del.message_id)

    message_for_del = await callback.message.answer(text='После установки пароля, данные о сотрудниках будут сохранены '
                                                         'на сервере, а Вам откроется функционал генерации "Секретных '
                                                         'кодов сотрудников", которые необходимы для регистрации '
                                                         'сотрудников в боте.')
    # "Функционал генерации" будет доступен из таблицы салари, где напротив каждой
    # записи будет значение кода руководителя - для формирования таблицы сотрудников
    # для которых конкретному руководителю доступна "генерация"
    small_message_for_delete.append(message_for_del.message_id)

    message_for_del = await callback.message.answer(
        text='Пароль должен состоять из английских букв и цифр, длинной от 7 до 10 символов'
    )
    small_message_for_delete.append(message_for_del.message_id)


def _check_salary_password(user_input: str) -> str | bool:
    password_pattern = re.compile(r'^(?=.*[A-Za-z])(?=.*\d)[A-Za-z\d]{7,10}$')

    if password_pattern.match(user_input):
        return user_input
    else:
        return False


@router.message(BossHere.creating_a_secret_code)
async def password_entry_processing(message: Message, state: FSMContext):
    """В случае корректности пароля, функция формирует dict_of_filling этой конкретной заливки
     и вместе с dict_of_persons передает в функцию insert_dict_of_persons_to_database"""
    global start_message_for_delete, small_message_for_delete

    # Помещаем ввод пользователя (с предполагаемым паролем в лист для удаления)
    small_message_for_delete.append(message.message_id)

    password = _check_salary_password(message.text)
    if password:
        current_datetime = datetime.now().strftime("%d.%m.%y %H:%M")
        author_of_entry = get_user_employee_code_from_db(str(message.chat.id))

        # Затираем статус "True" для всех имеющихся (старых) строк данного руководителя:
        update_data_in_column(table_name=TABLE_NAME,
                              base_column_name='author_of_entry',
                              base_column_value=author_of_entry,
                              target_column_name='available_to_supervisor',
                              new_value='False')

        # Формируем новые записи для данного руководителя и заливаем в БД:
        dict_of_filling = {'Report_card_date': str(current_datetime),
                           'author_of_entry': author_of_entry,
                           'salary_password': password,
                           'available_to_supervisor': 'True',
                           'available_to_employee': 'True'}
        insert_dict_of_persons_to_database(dict_of_persons, dict_of_filling)
        await message.answer(text=f"Пароль для табеля от {current_datetime} установлен: {password}")

        # Сбрасываем статус
        await state.set_state(Registration.employee_is_registered)

        # Удаляем сообщения, помеченные для удаления
        await delete_some_messages(chat_id=message.chat.id, numbers_of_message=small_message_for_delete)
        small_message_for_delete = []
        await delete_some_messages(chat_id=message.chat.id, numbers_of_message=start_message_for_delete)
        start_message_for_delete = []

        # Предлагаем сгенерировать секретные пароли сотрудникам
        message_for_del = await message.answer(text='Теперь Вы можете распечатать секретные пароли для сотрудников'
                                                    'из последней заливки табеля',
                                               reply_markup=make_inline_rows_keyboard(["Сгенерировать секретные коды"]))
        start_message_for_delete.append(message_for_del.message_id)

    else:
        message_for_del = await message.answer(text='Пароль не соответствует требованиям. Попробуйте еще раз')
        small_message_for_delete.append(message_for_del.message_id)
        message_for_del = await message.answer(text='Пароль должен состоять из английских букв и цифр,'
                                                    ' длинной от 7 до 10 символов')
        small_message_for_delete.append(message_for_del.message_id)


@router.callback_query(Registration.employee_is_registered, F.data.startswith("Сгенерировать секретные коды"))
async def starting_to_create_password_report(callback: CallbackQuery):
    await callback.answer()

    # Получаем код сотрудника руководителя:
    author_of_entry = get_user_employee_code_from_db(str(callback.message.chat.id))

    # Получаем все коды сотрудников для "данного руководителя" и "доступные руководителю":
    employee_code_of_autor = v_look_up_many(table_name=TABLE_NAME,
                                            base_column_names=['Author_of_entry', 'Available_to_supervisor'],
                                            base_column_values=[author_of_entry, 'True'],
                                            target_column_name='Employee_code')

    # Добавляем в список две дополнительные кнопки
    list_for_markup = ['Все сотрудники']
    list_for_markup.extend(employee_code_of_autor)
    list_for_markup.extend(['Генерация'])
    print(list_for_markup)

    await callback.message.answer(text='Выберете одного, несколько или всех сотрудников, '
                                       'а затем нажмите кнопку "Генерация"',
                                  reply_markup=make_inline_secret_many_keys_keyboard(list_for_markup))

# TODO необходима функция "выслать информацию", когда босс заливает ЗП, всем зарегистрированным должно придти уведомление

# TODO функция ловящая калбеки со списком сотрудников