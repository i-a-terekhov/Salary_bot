from aiogram import Router, F, types
from aiogram.filters import MagicData, CommandStart
from aiogram.types import Message, CallbackQuery
from aiogram.utils.keyboard import InlineKeyboardBuilder

from bot import bot_unit as bot
from hidden.tokenfile import OWNER_CHAT_ID

import openpyxl

# Хэндлеры этого роутера перехватят все сообщения и колбэки,
# если maintenance_mode равен True
maintenance_router = Router()
maintenance_router.message.filter(MagicData(F.maintenance_mode.is_(True)))
maintenance_router.callback_query.filter(MagicData(F.maintenance_mode.is_(True)))

regular_router = Router()


@maintenance_router.message()
async def any_message(message: Message):
    text = f'Пользователь {message.from_user.username} (ID={message.from_user.id}), заинтересовался ботом'
    print(text)
    await bot.send_message(chat_id=OWNER_CHAT_ID, text=text)
    await message.answer("Бот в режиме обслуживания. Пожалуйста, подождите.")


@maintenance_router.callback_query()
async def any_callback(callback: CallbackQuery):
    text = f'Пользователь {callback.from_user.username} (ID={callback.from_user.id}), заинтересовался ботом'
    print(text)
    await bot.send_message(chat_id=OWNER_CHAT_ID, text=text)
    await callback.answer(
        text="Бот в режиме обслуживания. Пожалуйста, подождите",
        show_alert=True
    )


# Хэндлеры этого роутера используются ВНЕ режима обслуживания,
# т.е. когда maintenance_mode равен False или не указан вообще
@regular_router.message(CommandStart())
async def cmd_start(message: Message):
    text = f'Пользователь {message.from_user.username} (ID={message.from_user.id}), заинтересовался ботом'
    print(text)
    await bot.send_message(chat_id=OWNER_CHAT_ID, text=text)
    builder = InlineKeyboardBuilder()
    builder.button(text="Нажми меня", callback_data="anything")
    await message.answer(
        text="Бот в разработке, т.е. пока не работает",
        reply_markup=builder.as_markup()
    )


@regular_router.message(F.document.file_name.endswith('.xlsx'))
async def handle_excel_file(message: types.Document):
    # Проверяем, что полученный файл - это Excel-файл
    if message.document.file_name.endswith('.xlsx'):
        # Скачиваем файл
        file_info = await bot.get_file(message.document.file_id)
        file = await bot.download_file(file_info.file_path)

        # Открываем файл с помощью openpyxl
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']

        # Извлекаем данные из диапазонов
        column_b_data = [cell[0].value for cell in sheet['B5:B11']]

        # Определение диапазона для извлечения данных
        start_cell = "C4"
        end_cell = "F5"
        # Получение численных значений для столбцов и строк
        start_col = openpyxl.utils.column_index_from_string(start_cell[:1])
        end_col = openpyxl.utils.column_index_from_string(end_cell[:1])
        start_row = int(start_cell[1:])
        end_row = int(end_cell[1:])

        row_c_data = []
        for row in range(start_row, end_row + 1):
            # Итерация по столбцам
            for col in range(start_col, end_col + 1):
                # Добавление значения в список
                row_c_data.append(sheet.cell(row=row, column=col).value)
        print(column_b_data)
        print(row_c_data)


        # Ищем ячейку со значением "Код сотрудника" в диапазоне A1:I17
        target_cell = None
        for row in sheet.iter_rows(min_row=1, max_row=17, min_col=1, max_col=9):
            for cell in row:
                if cell.value == "Код сотрудника":
                    target_cell = cell
                    break
            if target_cell:
                break

        # Если ячейка найдена, находим столбец "ЗП"
        zp_column = None
        if target_cell:
            for col in sheet.iter_cols(min_row=target_cell.row, max_row=sheet.max_row, min_col=target_cell.column,
                                       max_col=sheet.max_column):
                for cell in col:
                    if cell.value == "ЗП":
                        zp_column = col
                        break
                if zp_column:
                    break

        # Если столбец "ЗП" найден, выводим значения для непустых значений в столбце "Код сотрудника"
        if zp_column:
            for cell in zp_column:
                code_value = sheet.cell(row=cell.row, column=target_cell.column).value
                if code_value is not None:
                    zp_value = cell.value
                    print(f"Код сотрудника: {code_value}, ЗП: {zp_value}")


        # Выводим полученные данные
        await message.answer(f"Column B data: {column_b_data}")
        await message.answer(f"Row C data: {row_c_data}")

        # Закрываем файл
        wb.close()
    else:
        await message.answer("Пожалуйста, отправьте файл с расширением .xlsx")



@regular_router.message(F.document != None)
async def catch_document(message: Message):
    text = f'Пользователь {message.from_user.username} (ID={message.from_user.id}) прислал документ:'
    print(text)
    doc_name = message.document.file_name
    await bot.send_document(chat_id=OWNER_CHAT_ID, document=message.document.file_id, caption=message.caption)

    builder = InlineKeyboardBuilder()
    builder.button(text="Нажми меня", callback_data="anything")
    await message.answer(
        text="Бот в разработке, т.е. пока не работает",
        reply_markup=builder.as_markup()
    )


@regular_router.message()
async def catch_any_types(message: Message):
    text = f'Пользователь {message.from_user.username} (ID={message.from_user.id}) прислал прочее сообщение:'
    print(text)
    await bot.forward_message(chat_id=OWNER_CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)

    builder = InlineKeyboardBuilder()
    builder.button(text="Нажми меня", callback_data="anything")
    await message.answer(
        text="Бот в разработке, т.е. пока не работает",
        reply_markup=builder.as_markup())


@regular_router.callback_query(F.data == "anything")
async def callback_anything(callback: CallbackQuery):
    text = f'Пользователь {callback.from_user.username} (ID={callback.from_user.id}), заинтересовался ботом'
    print(text)
    await bot.send_message(chat_id=OWNER_CHAT_ID, text=text)
    await callback.answer(
        text="Кнопка, очевидно, пока тоже не работает :(",
        show_alert=True
    )
