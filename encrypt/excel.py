import os

import openpyxl
from aiogram.types import CallbackQuery, FSInputFile
from openpyxl.styles import Border, Side, Alignment

from database.general_db_functions import get_data_from_column
from encrypt.math_operations import check_employee_code
from handlers.salary import TABLE_NAME, bot


async def get_secret_codes_for_employee(callback: CallbackQuery, employee_codes: list | tuple) -> None:
    """Функция принимает список сотрудников и отправляет руководителю Excel-файл с секретными кодами для них"""

    # Создаем новый Excel-файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Создаем базовый лист со всеми выбранными сотрудниками:
    for code in employee_codes:
        name = get_data_from_column(table_name=TABLE_NAME,
                                    base_column_name='Employee_code',
                                    base_column_value=code,
                                    target_column_name='Full_name'
                                    )[0]
        secret = check_employee_code(code)
        # Добавляем данные в таблицу Excel
        sheet.append([code, name, secret])

    # Создаем стиль для ячеек с границами
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Рассчитываем количество строк на листе
    rows_per_page = 22  # Выведено экспериментальным путем
    max_rows = sheet.max_row

    # Создаем новые листы, учитывая ограничение строк на каждом
    for start_row_print in range(1, max_rows, rows_per_page):
        end_row_print = min(start_row_print + rows_per_page - 1, max_rows)
        new_sheet = workbook.create_sheet(title=f'Page_{(start_row_print - 1) // rows_per_page + 1}')
        new_sheet.append(["Код", "Ф.И.О.", "Secret Code", "Примечание"])

        # Копируем соответствующие строки данных
        for row in range(start_row_print, end_row_print + 1):
            new_sheet.append(
                [sheet.cell(row=row, column=1).value,
                 sheet.cell(row=row, column=2).value,
                 sheet.cell(row=row, column=3).value,
                 'Уничтожьте эту информацию!']
            )

        for row in new_sheet.iter_rows(min_row=1, max_col=new_sheet.max_column, max_row=new_sheet.max_row):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(vertical='center')
                    cell.border = thin_border
                    new_sheet.row_dimensions[cell.row].height = 30

        # Задаем ширину столбцов:
        new_sheet.column_dimensions['A'].width = 10
        new_sheet.column_dimensions['B'].width = 35
        new_sheet.column_dimensions['C'].width = 20
        new_sheet.column_dimensions['D'].width = 30
        new_sheet.page_setup.fitToHeight = 0
        new_sheet.page_setup.fitToWidth = 1
        new_sheet.page_setup.fitToPage = True
        new_sheet.page_setup.print_area = new_sheet.calculate_dimension()

    # Удаляем лист 'Sheet'
    sheet = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(sheet)

    # Определяем путь к файлу в директории бота
    file_path = "secret_codes.xlsx"

    # Сохраняем Excel-файл в директории бота
    workbook.save(file_path)

    # Оборачиваем файл в обертку FSInputFile и отправляем руководителю
    table = FSInputFile(file_path)
    await bot.send_document(chat_id=callback.message.chat.id, document=table)

    # Закрываем Excel-файл
    workbook.close()

    # Удаляем файл после отправки:
    os.remove(file_path)
